#!/usr/bin/env python3
"""Audit the current Sobag catalog and write human-editable reports.

The audit is read-only for the site catalog. Reports are written to
local-import-output/catalog-audit by default, which is ignored by git.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

from product_importer import product_variants


PRODUCT_WORDS = ["Подушка", "Наволочка", "Плед", "Мешок", "Чехол", "Флаг", "Ремувка"]
PLANNED_CATEGORIES = ["Подушки", "Наволочки", "Пледы", "Мешки для обуви", "Чехлы на кулер", "Чехлы на чемодан"]
RISKY_NAME_WORDS = ["Найк", "Nike", "Симпсоны", "Стрей Кидс", "Stray Kids"]
CANONICAL_ALIASES = {
    "однотонный": "Однотонные",
    "однотонная": "Однотонные",
    "однотонное": "Однотонные",
    "однотонные": "Однотонные",
    "детский": "Детские",
    "детская": "Детские",
    "детское": "Детские",
    "детские": "Детские",
    "женский": "Женские",
    "женская": "Женские",
    "женские": "Женские",
    "мужской": "Мужские",
    "мужская": "Мужские",
    "мужские": "Мужские",
    "животное": "Животные",
    "животные": "Животные",
    "паттерн": "Паттерны",
    "паттерны": "Паттерны",
    "подарок": "Подарки",
    "подарки": "Подарки",
    "мем": "Мемы",
    "мемы": "Мемы",
}

SHEET_FILE_NAMES = {
    "Сводка": "summary.csv",
    "Категории": "categories.csv",
    "Справочники": "taxonomy.csv",
    "Дубли справочников": "taxonomy-duplicates.csv",
    "Замечания": "data-quality.csv",
    "Названия": "naming-issues.csv",
    "Фото": "photo-report.csv",
    "Цены товаров": "product-price-summary.csv",
    "Цены вариантов": "variant-prices.csv",
}


def write_csv(path: Path, headers: list[str], rows: Iterable[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(path: Path, sheets: dict[str, tuple[list[str], list[dict[str, object]]]]) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for title, (headers, rows) in sheets.items():
        sheet = workbook.create_sheet(title[:31])
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, start=1):
            values = [str(header), *[str(row.get(header, "")) for row in rows[:200]]]
            width = min(max(len(value) for value in values) + 2, 60)
            sheet.column_dimensions[get_column_letter(index)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def load_products(path: Path) -> list[dict]:
    return json.loads(path.read_text(encoding="utf-8"))


def clean_key(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def canonical(value: str) -> str:
    key = clean_key(value)
    return CANONICAL_ALIASES.get(key, str(value or "").strip())


def list_values(product: dict, key: str) -> list[str]:
    value = product.get(key)
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if value:
        return [str(value).strip()]
    return []


def sample_skus(products: list[dict], limit: int = 8) -> str:
    return "; ".join(str(product.get("baseSku", "")) for product in products[:limit])


def category_rows(products: list[dict]) -> list[dict[str, object]]:
    by_category: dict[str, list[dict]] = defaultdict(list)
    for product in products:
        for category in list_values(product, "categories") or list_values(product, "category"):
            by_category[category].append(product)
    return [
        {
            "Категория": category,
            "Товаров": len(items),
            "Доля от каталога": f"{len(items) / len(products):.1%}",
            "Примеры артикулов": sample_skus(items),
        }
        for category, items in sorted(by_category.items(), key=lambda item: (-len(item[1]), item[0]))
    ]


def taxonomy_rows(products: list[dict]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for kind, key in [("Категория", "categories"), ("Подборка", "collections"), ("Праздник", "holidays"), ("Тег", "tags"), ("Тип", "types"), ("Размер", "sizes"), ("Материал", "materials")]:
        grouped: dict[str, list[dict]] = defaultdict(list)
        for product in products:
            for value in list_values(product, key):
                grouped[value].append(product)
        for value, items in sorted(grouped.items(), key=lambda item: (-len(item[1]), item[0])):
            suggestion = canonical(value)
            rows.append(
                {
                    "Тип справочника": kind,
                    "Значение": value,
                    "Товаров": len(items),
                    "Рекомендуемое значение": suggestion if suggestion != value else "",
                    "Примеры артикулов": sample_skus(items),
                }
            )
    return rows


def duplicate_taxonomy_rows(products: list[dict]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for kind, key in [("Подборка", "collections"), ("Тег", "tags"), ("Праздник", "holidays")]:
        grouped: dict[str, Counter] = defaultdict(Counter)
        examples: dict[str, list[str]] = defaultdict(list)
        for product in products:
            for value in list_values(product, key):
                prepared = canonical(value)
                grouped[prepared][value] += 1
                if len(examples[prepared]) < 8:
                    examples[prepared].append(str(product.get("baseSku", "")))
        for prepared, variants in grouped.items():
            if len(variants) > 1:
                rows.append(
                    {
                        "Тип справочника": kind,
                        "Рекомендуемое значение": prepared,
                        "Варианты в каталоге": "; ".join(sorted(variants)),
                        "Товаров": sum(variants.values()),
                        "Примеры артикулов": "; ".join(examples[prepared]),
                    }
                )
    return sorted(rows, key=lambda row: (row["Тип справочника"], row["Рекомендуемое значение"]))


def data_quality_rows(products: list[dict]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    category_counts = Counter(value for product in products for value in list_values(product, "categories"))
    for category in PLANNED_CATEGORIES:
        if category_counts[category] == 0:
            rows.append(
                {
                    "Важность": "medium",
                    "Раздел": "Категории",
                    "Проблема": "Плановая категория пока без товаров",
                    "Основной артикул": "",
                    "Значение": category,
                    "Что сделать": "Добавить товары в категорию или временно скрыть ее на витрине.",
                }
            )

    empty_collections = [product for product in products if not list_values(product, "collections")]
    if empty_collections:
        rows.append(
            {
                "Важность": "medium",
                "Раздел": "Подборки",
                "Проблема": "Есть товары без подборок",
                "Основной артикул": sample_skus(empty_collections, 20),
                "Значение": len(empty_collections),
                "Что сделать": "Заполнить подборки, если товар должен попадать в витринные разделы.",
            }
        )

    for product in products:
        sizes = list_values(product, "sizes")
        name = str(product.get("name", ""))
        if "13х3" in sizes and "5" in sizes:
            rows.append(
                {
                    "Важность": "high",
                    "Раздел": "Размеры",
                    "Проблема": "Размер похож на ошибочно разделенный десятичный формат",
                    "Основной артикул": product.get("baseSku", ""),
                    "Значение": "; ".join(sizes),
                    "Что сделать": "Проверить: вероятно, нужно одно значение `13х3,5`.",
                }
            )
        for word in RISKY_NAME_WORDS:
            if word.casefold() in name.casefold():
                rows.append(
                    {
                        "Важность": "medium",
                        "Раздел": "Названия",
                        "Проблема": "Название содержит бренд/известную тему, стоит проверить права",
                        "Основной артикул": product.get("baseSku", ""),
                        "Значение": name,
                        "Что сделать": "Перед публикацией проверить допустимость продажи/описания.",
                    }
                )
                break
    return rows


def repeated_product_word_issue(name: str) -> str:
    words = re.findall(r"[A-Za-zА-Яа-яЁё0-9]+", name)
    lowered = [word.casefold() for word in words]
    for index in range(len(lowered) - 1):
        if lowered[index] == lowered[index + 1] and any(lowered[index] == word.casefold() for word in PRODUCT_WORDS):
            return f"Повтор слова: {words[index]} {words[index + 1]}"
    for product_word in PRODUCT_WORDS:
        matches = [word for word in words[:4] if word.casefold() == product_word.casefold()]
        if len(matches) > 1:
            return f"Повтор товарного слова: {product_word}"
    return ""


def naming_issue_rows(products: list[dict]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    seen_names: dict[str, list[dict]] = defaultdict(list)
    for product in products:
        seen_names[clean_key(product.get("name", ""))].append(product)
        name = str(product.get("name", "")).strip()
        description = str(product.get("description", "")).strip()
        detail = str(product.get("detailDescription", "")).strip()
        issues = []
        if not name:
            issues.append(("critical", "Пустое название"))
        if len(name) > 70:
            issues.append(("medium", "Длинное название"))
        repeated = repeated_product_word_issue(name)
        if repeated:
            issues.append(("high", repeated))
        if name.startswith("Товар "):
            issues.append(("high", "Шаблонное название"))
        if not description or "Карточка импортирована" in description:
            issues.append(("medium", "Пустое или шаблонное краткое описание"))
        if not detail or "Карточка импортирована" in detail:
            issues.append(("medium", "Пустое или шаблонное описание в карточке"))
        for severity, issue in issues:
            rows.append(
                {
                    "Важность": severity,
                    "Проблема": issue,
                    "Основной артикул": product.get("baseSku", ""),
                    "Название": name,
                    "Краткое описание": description,
                    "Описание в карточке": detail,
                }
            )
    for normalized_name, items in seen_names.items():
        if normalized_name and len(items) > 1:
            rows.append(
                {
                    "Важность": "low",
                    "Проблема": "Одинаковое базовое название у нескольких товаров",
                    "Основной артикул": sample_skus(items, 20),
                    "Название": items[0].get("name", ""),
                    "Краткое описание": "",
                    "Описание в карточке": "",
                }
            )
    return rows


def image_info(path: Path) -> tuple[int, int]:
    with Image.open(path) as image:
        return image.size


def photo_rows(products: list[dict], root: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for product in products:
        images = [product.get("image"), *(product.get("gallery") or [])]
        images = [str(image) for image in images if image]
        issues = []
        width = height = ""
        if not images:
            issues.append("Нет фото")
        else:
            main_path = root / images[0]
            if not main_path.exists():
                issues.append("Главное фото не найдено")
            else:
                try:
                    width, height = image_info(main_path)
                    ratio = width / height if height else 0
                    if ratio < 0.9 or ratio > 1.1:
                        issues.append("Главное фото не квадратное")
                    if min(width, height) < 700:
                        issues.append("Главное фото меньше 700px по короткой стороне")
                except Exception as error:
                    issues.append(f"Не удалось прочитать главное фото: {error}")
        if len(images) < 2:
            issues.append("Мало фото: меньше 2")
        rows.append(
            {
                "Основной артикул": product.get("baseSku", ""),
                "Категории": "; ".join(list_values(product, "categories")),
                "Фото всего": len(images),
                "Главное фото": images[0] if images else "",
                "Ширина": width,
                "Высота": height,
                "Проблемы": "; ".join(issues),
            }
        )
    return rows


def variant_rows(products: list[dict]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for product in products:
        for variant in product_variants(product):
            rows.append(
                {
                    "Основной артикул": product.get("baseSku", ""),
                    "Артикул варианта": variant["sku"],
                    "Название варианта": variant["name"],
                    "Тип товара": variant["type"],
                    "Размер": variant["size"],
                    "Материал": variant["material"],
                    "Цена варианта": variant["price"],
                    "Категории": "; ".join(list_values(product, "categories")),
                    "Подборки": "; ".join(list_values(product, "collections")),
                    "Праздники": "; ".join(list_values(product, "holidays")),
                    "Теги": "; ".join(list_values(product, "tags")),
                    "Папка фото": product.get("photoFolder", product.get("baseSku", "")),
                }
            )
    return rows


def price_summary_rows(products: list[dict]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for product in products:
        variants = product_variants(product)
        prices = [int(variant["price"]) for variant in variants]
        rows.append(
            {
                "Основной артикул": product.get("baseSku", ""),
                "Название": product.get("name", ""),
                "Вариантов": len(variants),
                "Цена мин": min(prices) if prices else "",
                "Цена макс": max(prices) if prices else "",
                "Базовая цена": product.get("basePrice", ""),
                "Типы": "; ".join(list_values(product, "types")),
                "Размеры": "; ".join(list_values(product, "sizes")),
                "Материалы": "; ".join(list_values(product, "materials")),
            }
        )
    return rows


def summary_rows(products: list[dict], sheets: dict[str, tuple[list[str], list[dict[str, object]]]]) -> list[dict[str, object]]:
    variants = sum(len(product_variants(product)) for product in products)
    with_photo_issues = sum(1 for row in sheets["Фото"][1] if row["Проблемы"])
    return [
        {"Метрика": "Дата аудита", "Значение": datetime.now().strftime("%Y-%m-%d %H:%M")},
        {"Метрика": "Базовых товаров", "Значение": len(products)},
        {"Метрика": "Расчетных вариантов", "Значение": variants},
        {"Метрика": "Категорий", "Значение": len({value for product in products for value in list_values(product, "categories")})},
        {"Метрика": "Подборок", "Значение": len({value for product in products for value in list_values(product, "collections")})},
        {"Метрика": "Праздников", "Значение": len({value for product in products for value in list_values(product, "holidays")})},
        {"Метрика": "Проблем с названиями/описаниями", "Значение": len(sheets["Названия"][1])},
        {"Метрика": "Отдельных замечаний к данным", "Значение": len(sheets["Замечания"][1])},
        {"Метрика": "Потенциальных дублей справочников", "Значение": len(sheets["Дубли справочников"][1])},
        {"Метрика": "Товаров с замечаниями по фото", "Значение": with_photo_issues},
    ]


def write_markdown_summary(path: Path, summary: list[dict[str, object]], key_findings: list[str]) -> None:
    lines = ["# Catalog Audit Summary", ""]
    for row in summary:
        lines.append(f"- {row['Метрика']}: {row['Значение']}")
    lines.append("")
    lines.append("## Key Findings")
    lines.extend(f"- {item}" for item in key_findings)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Audit Sobag live catalog")
    parser.add_argument("--products", default="data/products-live.json", help="products JSON path")
    parser.add_argument("--out", default="local-import-output/catalog-audit", help="report output directory")
    args = parser.parse_args()

    root = Path.cwd()
    out_dir = root / args.out
    products = load_products(root / args.products)

    sheets: dict[str, tuple[list[str], list[dict[str, object]]]] = {
        "Категории": (["Категория", "Товаров", "Доля от каталога", "Примеры артикулов"], category_rows(products)),
        "Справочники": (["Тип справочника", "Значение", "Товаров", "Рекомендуемое значение", "Примеры артикулов"], taxonomy_rows(products)),
        "Дубли справочников": (["Тип справочника", "Рекомендуемое значение", "Варианты в каталоге", "Товаров", "Примеры артикулов"], duplicate_taxonomy_rows(products)),
        "Замечания": (["Важность", "Раздел", "Проблема", "Основной артикул", "Значение", "Что сделать"], data_quality_rows(products)),
        "Названия": (["Важность", "Проблема", "Основной артикул", "Название", "Краткое описание", "Описание в карточке"], naming_issue_rows(products)),
        "Фото": (["Основной артикул", "Категории", "Фото всего", "Главное фото", "Ширина", "Высота", "Проблемы"], photo_rows(products, root)),
        "Цены товаров": (["Основной артикул", "Название", "Вариантов", "Цена мин", "Цена макс", "Базовая цена", "Типы", "Размеры", "Материалы"], price_summary_rows(products)),
        "Цены вариантов": (["Основной артикул", "Артикул варианта", "Название варианта", "Тип товара", "Размер", "Материал", "Цена варианта", "Категории", "Подборки", "Праздники", "Теги", "Папка фото"], variant_rows(products)),
    }
    sheets["Сводка"] = (["Метрика", "Значение"], summary_rows(products, sheets))

    out_dir.mkdir(parents=True, exist_ok=True)
    for name, (headers, rows) in sheets.items():
        write_csv(out_dir / SHEET_FILE_NAMES.get(name, f"{name}.csv"), headers, rows)
    write_workbook(out_dir / "catalog-audit.xlsx", sheets)

    key_findings = [
        f"Каталог содержит {len(products)} базовых товаров и {len(sheets['Цены вариантов'][1])} расчетных вариантов.",
        f"Замечаний по фото: {sum(1 for row in sheets['Фото'][1] if row['Проблемы'])}.",
        f"Замечаний по названиям/описаниям: {len(sheets['Названия'][1])}.",
        f"Отдельных замечаний к данным: {len(sheets['Замечания'][1])}.",
        f"Потенциальных дублей справочников: {len(sheets['Дубли справочников'][1])}.",
    ]
    write_markdown_summary(out_dir / "summary.md", sheets["Сводка"][1], key_findings)
    print(json.dumps({"out": str(out_dir), "products": len(products), "variants": len(sheets["Цены вариантов"][1])}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
